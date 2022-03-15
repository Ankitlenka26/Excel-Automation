import openpyxl

wb = openpyxl.load_workbook("input.xlsx")
sheetObj = wb["Sheet1"]

print("Python defined max_column " + str(sheetObj.max_column))
print("Python defined max_row " + str(sheetObj.max_row))


def get_maximum_cols():
    for i in range(1, 20000):
        if sheetObj.cell(row=2, column=i).value == None:
            max_col = i
            break
    return max_col


def get_maximum_rows():
    for i in range(1, 20000):
        if sheetObj.cell(row=i, column=2).value == None:
            max_row = i
            break
    return max_row


row = get_maximum_rows()
column = get_maximum_cols()
print('max column ' + str(column))
print('max row ' + str(row))
wb.save("input.xlsx")

# sheetObj = wb['Sheet1']
gameIds = [0 for i in range(row-4)]  # present in 1st column
noOfGamesPerBoot = [[0 for i in range(column-2)] for j in range(row-4)]
bootPrice = [0 for i in range(column-2)]
bootDurations = [0 for i in range(column-2)]
# populating the noOfGamesPerBoot matrix
for i in range(row-4):
    for j in range(column-2):
        noOfGamesPerBoot[i][j] = int(sheetObj.cell(i+4, j+2).value)

# for adding gameIds
for i in range(row-4):
    gameIds[i] = int(sheetObj.cell(i+4, 1).value)

# for adding bootDurations
for i in range(column-2):
    bootDurations[i] = int(sheetObj.cell(2, i+2).value)
    bootPrice[i] = int(sheetObj.cell(1, i+2).value)

# print(gameIds)
# print(bootPrice)
# print(bootDurations)
# print(noOfGamesPerBoot)
