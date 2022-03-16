from openpyxl import Workbook


def outputMainFunc(finalData, noOfBoots, noOfGames):
    wb = Workbook()
    sheetObj = wb.active

    # print the first row "the names"
    sheetObj.cell(row=1, column=1, value='entry_fee')
    sheetObj.cell(row=1, column=2, value='start_time')
    sheetObj.cell(row=1, column=3, value='end_time')
    sheetObj.cell(row=1, column=4, value='game_id')
    # print the data now to excel
    cnt = 2
    for i in range(noOfBoots):
        for j in range(noOfGames[i]):
            sheetObj.cell(row=cnt, column=1, value=finalData[i][j][0])
            sheetObj.cell(row=cnt, column=2, value=finalData[i][j][1])
            sheetObj.cell(row=cnt, column=3, value=finalData[i][j][2])
            sheetObj.cell(row=cnt, column=4, value=finalData[i][j][3])
            cnt += 1

    wb.save("output.xlsx")
    print('Output is now present in excel')
    return None
